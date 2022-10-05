from openpyxl import load_workbook


def write_excel(r, c, v):
    v = str(v)
    wb = load_workbook('TA_TEST.xlsx')
    sh = wb.active
    cell = sh.cell(row=r, column=c)
    cell.value = v
    print(f'chosen row = {r}, col = {c}')
    print('value added to excel:', cell.value)
    wb.save('TA_TEST.xlsx')
    #input("Save complete")


if __name__ == '__main__':
    pass
